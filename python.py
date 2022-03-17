import openpyxl
from openpyxl import Workbook
wb = Workbook()
sh1 = wb.active
def fun1(register):
    username = []
    while register == "yes":
        u = input("  Enter your email id to register: ")
        x = list(map(str, u))
        if "@" in x and "." in x and x[0].isalpha():
            y = u.split("@")
            z = list(map(str, y[1]))
            l = y[1].split(".")
            if l[0].isalpha() or l[0].isalnum():
                username.append(u)
                break
            else:
                print("-------Enter a your email id correctly------ ")

        else:
            print("-------Enter a your email id correctly------ ")
            continue
    sh1.cell(row=1, column=1).value = username[0]
    wb.save("registration.xlsx")
    return username
def fun2(register):
    password = []
    while register == "yes":
        print("\n")
        print("please check the below points before setting up your password")
        print("\n")
        print("1. Must have minimum one special character")
        print("2. Must have minimum one digit")
        print("3. Must have minimum one uppercase")
        print("4. Must have minimum one lowercase")
        print("5. Password length must be between 5 and 16")
        print("\n")
        u = input("  Enter your password to register: ")
        x = list(map(str, u))
        w=len(x)
        lower = 0
        upper = 0
        num=0
        sp=0
        for i in x:
            if i.isupper():
                upper += 1
            elif i.islower():
                lower+= 1
            elif i.isdigit():
                num+=1
            else:
                sp+=1
        if w>5 and w<16 and lower!=0 and upper != 0 and sp!=0 and num!=0:
            password.append(u)
            break
        else:
            print("-------Enter a your password correctly------ ")
            continue
    sh1.cell(row=1, column=2).value = password[0]
    wb.save("registration.xlsx")
    return password



register=input(" Do you want to register?   yes or no :")
if register=="yes":
    username = fun1(register)
    password = fun2(register)


login=input(" Do you want to login yes or no :")
if login=="yes":
    a=input("enter username")
    b=input("enter password")
    wb1=openpyxl.load_workbook("registration.xlsx")
    sh2=wb1.active
    data1=sh2.cell(row=1, column=1).value
    data2=sh2.cell(row=1, column=2).value
    if a==data1 and b==data2:
        print("you have logged in sucessfully")
    elif a!=data1 and b!=data2:
        print("Invaild username or password")
        print(" Try to register before loging in")
        x = input("Do you want to register? yes or no :")
        if x == "yes":
            fun1(x)
            fun2(x)
elif login=="no":
    e = input("forget password yes or no :")
    if e == "yes":
        f = input("enter your username :")
        wb3 = openpyxl.load_workbook("registration.xlsx")
        sh3 = wb3.active
        data11 = sh3.cell(row=1, column=1).value
        data22 = sh3.cell(row=1, column=2).value
        if f==data11:
            print("your password is:",data22)
        else:
            print("-----sorry Wrong username-----")
